#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
골드셋 수동 라벨링 시트 생성기 (FNC / 엔플라잉 Gap 분석)

각 팀원이 '자기 band_all.jsonl'에 대해 한 번씩 실행한다.
  - band_all 에서 N개(기본 150)를 뽑아
  - 사람이 바로 라벨링할 수 있는 .xlsx 파일을 만든다.
  - 실행하면 이상 여부(파싱 실패/중복/비한국어/짧은 댓글 등)를 리포트로 출력한다.

사용 예:
  python3 make_labeling_sheet.py --input band_all.jsonl --name 수빈
  python3 make_labeling_sheet.py --input data/train/band/juhyeong/band_all.jsonl --name juhyeong

필요 패키지:  pip install openpyxl
"""

import argparse
import json
import os
import random
import re
import sys
from collections import Counter, OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    sys.exit("openpyxl 이 필요합니다.  먼저 실행:  pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# 라벨 정의 (SSOT). 순서 = 시트 컬럼 순서.
# ─────────────────────────────────────────────────────────────────────────────
LABELS = [
    "밴드_에너지", "연주_악기", "보컬_라이브", "이별_감성", "위로_공감",
    "청량_여름", "신규_유입", "장기_팬덤", "역주행_기대", "기타_노이즈",
]

LABEL_GUIDE = [
    ("밴드_에너지", "밴드 정체성·락 사운드·돌파 에너지에 대한 반응", '"밴드 느낌난다", "락 에너지", "합주 미쳤다"'),
    ("연주_악기",  "악기·연주(드럼/베이스/기타 등) 자체에 대한 반응",   '"드럼 미쳤다", "베이스 라인", "기타 솔로 소름"'),
    ("보컬_라이브", "가창력·음색·라이브 보컬에 대한 반응",            '"라이브 소름", "음색 독보적", "고음 미쳤다"'),
    ("이별_감성",  "이별·슬픔·먹먹한 정서",                        '"이별 공감", "드라마 OST 느낌", "울었다"'),
    ("위로_공감",  "위로·공감·내 얘기 같다는 반응",                 '"내 얘기 같다", "위로된다", "괜찮아질 것 같다"'),
    ("청량_여름",  "밝고 시원한·청량한 에너지",                     '"청량하다", "여름 노래", "기분 좋아진다"'),
    ("신규_유입",  "새로 유입된 정황(드라마·알고리즘·처음 알게 됨)",   '"드라마 보고 왔다", "처음 알았는데", "알고리즘 떠서"'),
    ("장기_팬덤",  "오래 봐온 팬의 정황",                          '"몇 년째 듣는다", "데뷔 때부터", "옛날 생각난다"'),
    ("역주행_기대", "역주행·숨은 명곡이라는 반응",                   '"역주행 가자", "숨겨진 명곡", "왜 안 뜨지"'),
    ("기타_노이즈", "위 어디에도 안 맞는 응원·잡담·이모지·단순 좋아요",  '"사랑해", "ㅋㅋㅋ", "1등", "❤️"'),
]

RULES = [
    "① 멀티라벨: 해당하는 라벨 칸에 모두 1. 한 댓글에 여러 개 가능.",
    "② 의미로 판단: 표면 단어가 없어도 맥락상 해당하면 1. (예: \"0:56 극락\" → 그 장면 몰입)",
    "③ 억지로 끼우지 말 것: 아무 데도 안 맞으면 '기타_노이즈' 하나만.",
    "④ 모르겠거나 새 라벨이 필요한 댓글 → 맨 오른쪽 'suggested' 칸에 자유 메모.",
]

EXAMPLES = [
    ('"데뷔 때부터 봤는데 라이브 진짜 미쳤다"', "장기_팬덤 + 보컬_라이브"),
    ('"드라마 보고 왔어요 ㅠㅠ 들으면서 울었다"', "신규_유입 + 이별_감성"),
    ('"역주행 가즈아 이건 숨은 명곡임"', "역주행_기대"),
    ('"ㅋㅋㅋㅋ 1등"', "기타_노이즈"),
]

REQUIRED_FIELDS = ["comment_id", "text", "video_type", "artist"]
HANGUL_RE = re.compile(r"[가-힣ㄱ-ㅎㅏ-ㅣ]")

# 색상
C_CTX_HEADER = "374151"   # 진회색 - 맥락 컬럼 헤더
C_LBL_HEADER = "1D4ED8"   # 파랑 - 라벨 컬럼 헤더
C_SUG_HEADER = "B45309"   # 호박색 - suggested 헤더
C_KEY_FILL   = "F3F4F6"   # 연회색 - comment_id 등 키 컬럼(수정금지 표시)
C_BAND_FILL  = "F9FAFB"   # 줄무늬
C_FILLED     = "DCFCE7"   # 라벨 입력 시 연초록
C_TITLE      = "111827"

FONT = "Arial"
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean_text(s):
    if s is None:
        return ""
    s = str(s)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)   # 엑셀이 거부하는 제어문자 제거
    return s


def text_row_height(s):
    """text 컬럼(너비~60)에서 줄바꿈/길이 기준 행 높이 추정 → 가독성."""
    if not s:
        return 22
    lines = 0
    for seg in s.split("\n"):
        w = sum(2 if ord(ch) > 0x1100 else 1 for ch in seg)  # 한글/이모지 ≈ 2배 폭
        lines += max(1, -(-w // 58))                          # ceil
    return max(22, min(lines, 6) * 15)                        # 최대 6줄


def load_pool(path, min_len, keep_non_korean):
    """JSONL 로드 + 정제. (정제된 레코드 리스트, 리포트 dict) 반환."""
    rep = Counter()
    raw = []
    if not os.path.exists(path):
        sys.exit("입력 파일을 찾을 수 없습니다: %s" % path)

    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rep["read"] += 1
            try:
                d = json.loads(line)
            except Exception:
                rep["bad_json"] += 1
                continue
            if not all(k in d for k in REQUIRED_FIELDS):
                rep["missing_field"] += 1
                continue
            raw.append(d)

    # 중복 제거: comment_id → 텍스트
    seen_id, seen_txt, out = set(), set(), []
    for d in raw:
        cid = str(d.get("comment_id", ""))
        if cid and cid in seen_id:
            rep["dup_id"] += 1
            continue
        seen_id.add(cid)

        txt = clean_text(d.get("text", "")).strip()
        norm = re.sub(r"\s+", " ", txt).lower()

        if len(txt) < min_len:
            rep["too_short"] += 1
            continue
        if not keep_non_korean and not HANGUL_RE.search(txt):
            rep["non_korean"] += 1
            continue
        if norm in seen_txt:
            rep["dup_text"] += 1
            continue
        seen_txt.add(norm)

        d["text"] = txt
        out.append(d)

    rep["pool"] = len(out)
    return out, rep


def sample(pool, n, seed, tilt):
    rng = random.Random(seed)
    if len(pool) <= n:
        picked = pool[:]
        rng.shuffle(picked)
        return picked

    if not tilt:
        return rng.sample(pool, n)

    # video_type 균형 보정: 타입별로 최대한 고르게 → 부족분은 랜덤 충원
    by_type = OrderedDict()
    for d in pool:
        by_type.setdefault(d.get("video_type", "?"), []).append(d)
    for v in by_type.values():
        rng.shuffle(v)

    per = max(1, n // len(by_type))
    picked, leftover = [], []
    for v in by_type.values():
        picked.extend(v[:per])
        leftover.extend(v[per:])
    rng.shuffle(leftover)
    picked.extend(leftover[: max(0, n - len(picked))])
    rng.shuffle(picked)
    return picked[:n]


# ─────────────────────────────────────────────────────────────────────────────
# 시트 작성
# ─────────────────────────────────────────────────────────────────────────────
def style_header(cell, fill_hex):
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def build_label_sheet(wb, rows, name):
    ws = wb.active
    ws.title = "라벨링"

    ctx_cols = ["번호", "comment_id", "artist", "video_type", "video_title", "text"]
    headers = ctx_cols + LABELS + ["suggested"]
    n_ctx = len(ctx_cols)                      # 6
    first_lbl = n_ctx + 1                      # 7 (G)
    last_lbl = n_ctx + len(LABELS)             # 16 (P)
    sug_col = last_lbl + 1                     # 17 (Q)

    # 헤더
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        if j <= n_ctx:
            style_header(c, C_CTX_HEADER)
        elif j <= last_lbl:
            style_header(c, C_LBL_HEADER)
        else:
            style_header(c, C_SUG_HEADER)

    # 데이터
    for i, d in enumerate(rows, start=1):
        r = i + 1
        band = (i % 2 == 0)
        vals = [
            i,
            clean_text(d.get("comment_id", "")),
            clean_text(d.get("artist", "")),
            clean_text(d.get("video_type", "")),
            clean_text(d.get("video_title", "")),
            clean_text(d.get("text", "")),
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            c.font = Font(name=FONT, size=10)
            if j == 6:   # text
                c.alignment = Alignment(wrap_text=True, vertical="center")
            elif j == 2:  # comment_id (키)
                c.alignment = Alignment(vertical="center")
                c.fill = PatternFill("solid", fgColor=C_KEY_FILL)
                c.font = Font(name=FONT, size=8, color="6B7280")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if band and j != 2:
                c.fill = PatternFill("solid", fgColor=C_BAND_FILL)

        # 라벨 + suggested 입력 칸
        for j in range(first_lbl, sug_col + 1):
            c = ws.cell(row=r, column=j)
            c.border = BORDER
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[r].height = text_row_height(vals[5])

    n = len(rows)
    last_row = n + 1

    # 컬럼 너비
    widths = {1: 5, 2: 22, 3: 10, 4: 10, 5: 30, 6: 60}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for col in range(first_lbl, last_lbl + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7.5
    ws.column_dimensions[get_column_letter(sug_col)].width = 22
    ws.row_dimensions[1].height = 50

    # 틀 고정: 1행(헤더) + A~F(맥락) 고정 → 스크롤해도 댓글/헤더가 보임
    ws.freeze_panes = "G2"

    # 라벨 칸 입력값 검증(드롭다운 '1'), 정보 스타일이라 막지 않음
    lbl_range = "%s2:%s%d" % (get_column_letter(first_lbl), get_column_letter(last_lbl), last_row)
    dv = DataValidation(type="list", formula1='"1"', allow_blank=True,
                        showErrorMessage=False, showInputMessage=False)  # 드롭다운만, 팝업 없음
    ws.add_data_validation(dv)
    dv.add(lbl_range)

    # 라벨 입력 시 초록색 강조(조건부 서식)
    green = PatternFill("solid", fgColor=C_FILLED)
    ws.conditional_formatting.add(
        lbl_range,
        FormulaRule(formula=["NOT(ISBLANK(%s2))" % get_column_letter(first_lbl)], fill=green),
    )

    ws.sheet_view.showGridLines = False
    return ws


def build_guide_sheet(wb, name, src, n, seed, rep):
    ws = wb.create_sheet("가이드(먼저읽기)", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 44

    def put(r, a, b=None, c=None, bold=False, size=10, fill=None, color="111827"):
        cell = ws.cell(row=r, column=1, value=a)
        cell.font = Font(name=FONT, bold=bold, size=size, color=color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill)
        if b is not None:
            cb = ws.cell(row=r, column=2, value=b)
            cb.font = Font(name=FONT, size=size)
            cb.alignment = Alignment(vertical="center", wrap_text=True)
        if c is not None:
            cc = ws.cell(row=r, column=3, value=c)
            cc.font = Font(name=FONT, size=size, color="6B7280")
            cc.alignment = Alignment(vertical="center", wrap_text=True)

    r = 1
    title = ws.cell(row=r, column=1, value="라벨링 가이드 — 먼저 읽어주세요")
    title.font = Font(name=FONT, bold=True, size=15, color=C_TITLE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 26
    r += 2

    put(r, "라벨러", name, bold=True); r += 1
    put(r, "원본 파일", src); r += 1
    put(r, "표본 수 / seed", "%d개 / seed=%d (재현 가능)" % (n, seed)); r += 1
    put(r, "정제 후 풀", "%d개에서 추출" % rep.get("pool", 0)); r += 2

    put(r, "작업 방법", bold=True, size=11, fill="E5E7EB"); r += 1
    put(r, "1.", "'라벨링' 탭으로 이동.  각 행(댓글)을 읽고, 해당하는 라벨 칸에 1 을 입력."); r += 1
    put(r, "2.", "한 댓글에 여러 라벨 가능(멀티라벨). 입력하면 칸이 초록색으로 표시됨."); r += 1
    put(r, "3.", "comment_id 열(회색)은 절대 수정하지 마세요. 나중에 원본과 합치는 키입니다."); r += 1
    put(r, "4.", "다 하면 파일 그대로 저장해서 주형에게 전달."); r += 2

    put(r, "라벨", "정의 / 판단 기준", "예시", bold=True, size=11, fill="DBEAFE"); r += 1
    for lab, desc, ex in LABEL_GUIDE:
        put(r, lab, desc, ex, bold=True if False else False)
        ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10, color="1D4ED8")
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1

    put(r, "규칙", bold=True, size=11, fill="E5E7EB"); r += 1
    for rule in RULES:
        put(r, rule); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 20
        r += 1
    r += 1

    put(r, "멀티라벨 예시", "→ 부여할 라벨", bold=True, size=11, fill="DBEAFE"); r += 1
    for ex, labs in EXAMPLES:
        put(r, ex, labs)
        r += 1

    return ws


def _print_report(out, name, rows, rep, min_len, n):
    art = Counter(d.get("artist", "?") for d in rows)
    vt = Counter(d.get("video_type", "?") for d in rows)
    lens = sorted(len(d["text"]) for d in rows)
    med = lens[len(lens) // 2] if lens else 0
    print("=" * 56)
    print(" 라벨링 시트 생성 완료:  %s" % out)
    print("=" * 56)
    print(" 라벨러            : %s" % name)
    print(" [정제 리포트]")
    print("   읽은 줄          : %d" % rep.get("read", 0))
    print("   JSON 파싱 실패   : %d" % rep.get("bad_json", 0))
    print("   필수필드 누락    : %d" % rep.get("missing_field", 0))
    print("   comment_id 중복  : %d" % rep.get("dup_id", 0))
    print("   텍스트 중복      : %d" % rep.get("dup_text", 0))
    print("   너무 짧음(<%d)    : %d" % (min_len, rep.get("too_short", 0)))
    print("   비한국어 제외    : %d%s" % (rep.get("non_korean", 0),
          "  (keep_non_korean=True 로 포함 가능)" if rep.get("non_korean", 0) else ""))
    print("   → 정제 후 풀     : %d" % rep.get("pool", 0))
    print(" [추출 표본 = %d개]" % len(rows))
    print("   아티스트 분포    : %s" % dict(art))
    print("   video_type 분포  : %s" % dict(vt))
    print("   텍스트 길이 중앙값: %d자" % med)
    if rep.get("pool", 0) < n:
        print("   ⚠ 풀이 요청 수보다 작아 전량을 사용했습니다.")
    print("=" * 56)


def generate(input_path, name=None, out=None, n=150, seed=42,
             tilt=False, keep_non_korean=False, min_len=2, verbose=True):
    """band_all.jsonl → 라벨링용 xlsx 생성. (출력경로, 리포트) 반환. 코랩/CLI 공용."""
    name = name or os.path.basename(os.path.dirname(os.path.abspath(input_path))) or "labeler"
    out = out or os.path.join(os.path.dirname(os.path.abspath(input_path)),
                              "%s_labeling_sheet.xlsx" % name)
    pool, rep = load_pool(input_path, min_len, keep_non_korean)
    if not pool:
        raise SystemExit("정제 후 남은 댓글이 없습니다. 입력/옵션을 확인하세요.")
    rows = sample(pool, n, seed, tilt)
    wb = Workbook()
    build_label_sheet(wb, rows, name)
    build_guide_sheet(wb, name, os.path.basename(input_path), len(rows), seed, rep)
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    if verbose:
        _print_report(out, name, rows, rep, min_len, n)
    return out, rep


# ======================== CLI (코랩에선 사용 안 함) ========================
def main():
    ap = argparse.ArgumentParser(description="골드셋 수동 라벨링 시트 생성기")
    ap.add_argument("--input", "-i", default="band_all.jsonl", help="band_all.jsonl 경로")
    ap.add_argument("--name", "-N", default=None, help="라벨러 이름(미지정 시 입력파일 상위 폴더명)")
    ap.add_argument("--n", type=int, default=150, help="추출 개수(기본 150)")
    ap.add_argument("--seed", type=int, default=42, help="랜덤 시드(기본 42)")
    ap.add_argument("--out", "-o", default=None, help="출력 xlsx 경로")
    ap.add_argument("--tilt", action="store_true", help="video_type 균형 보정 추출(기본: 순수 랜덤)")
    ap.add_argument("--keep-non-korean", action="store_true", help="한글 없는 댓글도 포함")
    ap.add_argument("--min-len", type=int, default=2, help="최소 글자수(기본 2)")
    args = ap.parse_args()
    generate(args.input, args.name, args.out, args.n, args.seed,
             args.tilt, args.keep_non_korean, args.min_len)


if __name__ == "__main__":
    main()
