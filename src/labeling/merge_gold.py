#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""골드 라벨링 시트(팀원별) → gold_labeled.jsonl 병합 + 라벨 분포 비교 리포트.

  python3 src/labeling/merge_gold.py [gold_dir]

각 시트의 'comment_id/artist/video_type/video_title/text + 10라벨 + suggested'를 읽어
멀티라벨 배열로 합치고, labeled_by(파일명)·label_source=human 을 붙인다.
"""
import glob
import json
import os
import sys
from collections import Counter, OrderedDict

from openpyxl import load_workbook

LABELS = ["밴드_에너지", "연주_악기", "보컬_라이브", "이별_감성", "위로_공감",
          "청량_여름", "신규_유입", "장기_팬덤", "역주행_기대", "기타_노이즈"]
CONCEPT = LABELS[:6]   # 의도 축 후보(컨셉/사운드)
FANDOM = LABELS[6:9]   # 신규·장기·역주행


def labeler_from(path):
    return os.path.basename(path).replace("_labeling_sheet.xlsx", "").replace(".xlsx", "")


def read_sheet(path):
    wb = load_workbook(path, read_only=True)
    ws = wb["라벨링"] if "라벨링" in wb.sheetnames else wb[wb.sheetnames[-1]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {name: i for i, name in enumerate(header) if name}

    def cell(r, k):
        i = col.get(k)
        return r[i] if i is not None and i < len(r) else None

    out = []
    for r in rows[1:]:
        if r is None:
            continue
        cid, txt = cell(r, "comment_id"), cell(r, "text")
        if (cid is None or str(cid).strip() == "") and (txt is None or str(txt).strip() == ""):
            continue
        labs = []
        for lab in LABELS:
            i = col.get(lab)
            if i is not None and i < len(r) and r[i] not in (None, "") and str(r[i]).strip() != "":
                labs.append(lab)
        out.append({
            "comment_id": cid, "artist": cell(r, "artist"),
            "video_type": cell(r, "video_type"), "video_title": cell(r, "video_title"),
            "text": txt, "label": labs, "suggested": cell(r, "suggested"),
        })
    return out


def main():
    gold_dir = sys.argv[1] if len(sys.argv) > 1 else "data/train/band/gold"
    out_path = os.path.join(gold_dir, "gold_labeled.jsonl")
    files = sorted(f for f in glob.glob(os.path.join(gold_dir, "*.xlsx"))
                   if "labeling_sheet" in os.path.basename(f))
    if not files:
        sys.exit("라벨 시트가 없습니다: %s" % gold_dir)

    all_recs, by = [], OrderedDict()
    for f in files:
        who = labeler_from(f)
        recs = read_sheet(f)
        for rec in recs:
            rec["labeled_by"] = who
            rec["label_source"] = "human"
        by[who] = recs
        all_recs.extend(recs)

    with open(out_path, "w", encoding="utf-8") as w:
        for rec in all_recs:
            w.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # ── 분포 비교 리포트 ──────────────────────────────────────────────
    per = {who: Counter() for who in by}
    tot = Counter()
    for who, recs in by.items():
        for rec in recs:
            for lab in rec["label"]:
                per[who][lab] += 1
                tot[lab] += 1

    print("=" * 70)
    print(" 병합: %s  (%d명, %d개)" % (out_path, len(by), len(all_recs)))
    print("=" * 70)
    namw = 12
    head = "라벨".ljust(namw) + "".join((w[:7]).rjust(9) for w in by) + "전체".rjust(9)
    print(head)
    for lab in LABELS:
        line = lab.ljust(namw) + "".join(str(per[who][lab]).rjust(9) for who in by) + str(tot[lab]).rjust(9)
        print(line)
    print("-" * 70)
    for who, recs in by.items():
        n = len(recs)
        etc = sum(1 for r in recs if r["label"] == ["기타_노이즈"])
        blank = sum(1 for r in recs if not r["label"])
        multi = sum(1 for r in recs if len(r["label"]) >= 2)
        artists = ",".join(sorted({str(r["artist"]) for r in recs}))
        print(" %-8s n=%3d  기타단독=%3d(%4.1f%%)  무라벨=%d  멀티=%d  [%s]"
              % (who, n, etc, 100 * etc / n if n else 0, blank, multi, artists))
    print("-" * 70)
    print(" 컨셉/사운드6 총 부여=%d  /  기타_노이즈=%d  /  팬덤(신규·장기·역주행)=%d"
          % (sum(tot[l] for l in CONCEPT), tot["기타_노이즈"], sum(tot[l] for l in FANDOM)))
    print("=" * 70)


if __name__ == "__main__":
    main()
